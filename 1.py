import openpyxl

wb = openpyxl.load_workbook('store.xlsx')
print(wb.sheetnames)

sheet = wb['Products']

# print(dir(sheet))

print(sheet.title)
# sheet.title = 'Products New'
# print(sheet.title)

# print(sheet.sheet_format)
# print(sheet.sheet_properties)

#wb.create_sheet('NewSheet1', 0)

# print(wb.sheetnames)
# sheet_to_remove = wb['NewSheet11']
# wb.remove(sheet_to_remove)
# print(wb.sheetnames)

# source = wb['NewSheet1']
# destination = wb.copy_worksheet(source)
# print(destination.title)


wb.save('store.xlsx')


